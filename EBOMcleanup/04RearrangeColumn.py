import os
from openpyxl import load_workbook

def move_columns(input_directory, output_directory):
    if input_directory == output_directory:
        print("Error: Input and output directories should be different.")
        return

    for root, dirs, files in os.walk(input_directory):
        for filename in files:
            if filename.endswith(".xlsx"):
                input_file_path = os.path.join(root, filename)
                output_file_path = input_file_path.replace(input_directory, output_directory)

                # Create the output directory if it doesn't exist
                os.makedirs(os.path.dirname(output_file_path), exist_ok=True)

                print(f"Processing file: {input_file_path}")

                # Load the workbook
                wb = load_workbook(input_file_path)
                ws = wb.active

                # Move columns C, D, E to H, I, J
                for row in range(1, ws.max_row + 1):
                    for src_col, dest_col in zip(range(3, 6), range(8, 11)):
                        ws.cell(row=row, column=dest_col).value = ws.cell(row=row, column=src_col).value

                # Delete original columns C, D, E
                for col in range(3, 6):
                    ws.delete_cols(3)

                # Save the modified workbook
                wb.save(output_file_path)
                print("Processed successfully")

# Prompt the user to enter the input and output directory paths
input_directory = input("Enter the input folder directory path: ")
output_directory = input("Enter the output folder directory path: ")

move_columns(input_directory, output_directory)
