import os
from openpyxl import Workbook, load_workbook

def combine_workbooks(input_directory, output_directory):
    if input_directory == output_directory:
        print("Error: Input and output directories should be different.")
        return

    for root, dirs, files in os.walk(input_directory):
        # Initialize a list to store data from all sheets
        all_data = []

        for filename in files:
            if filename.endswith(".xlsx"):
                input_file_path = os.path.join(root, filename)

                # Load the workbook
                wb = load_workbook(input_file_path)

                # Iterate through each sheet in the workbook
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]

                    # Extract data from each row and append to all_data list
                    for row in sheet.iter_rows(values_only=True):
                        all_data.append(row)

        # Create a new workbook for combined data
        output_filename = files[0][:12] + ".xlsx"
        output_file_path = os.path.join(output_directory, output_filename)
        new_wb = Workbook()

        # Create a new sheet in the new workbook
        new_sheet = new_wb.active
        new_sheet.title = "Combined_Data"

        # Write data to the new sheet
        for row_idx, row_data in enumerate(all_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                new_sheet.cell(row=row_idx, column=col_idx, value=value)

        # Save the new workbook
        new_wb.save(output_file_path)
        print(f"Combined data saved to: {output_file_path}")

# Prompt the user to enter the input and output directory paths
input_directory = input("Enter the input folder directory path: ")
output_directory = input("Enter the output folder directory path: ")

combine_workbooks(input_directory, output_directory)
